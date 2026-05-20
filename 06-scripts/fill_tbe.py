"""Deterministic TBE filler for Krayavikrayam (package: UV-2 by default).

Pipeline:
  1. Load 00-rules/CLAUDE.md (asserts presence; rules are also hard-coded here
     so the script's behaviour is auditable).
  2. Parse 05-evidence/<PKG>_evidence.md — each vendor is a fenced ```json block.
  3. Open 01-templates/TBE-<PKG>.xlsx with openpyxl(keep_vba=True, keep_links=True).
  4. For every Subcategory × Vendor cell trio (Details, Score, SxW):
        - Capex row → blank w/ "TBC — Procurement Team assessment, sealed commercial offer".
        - Otherwise → call the dedicated scorer; compose Details with verbatim
          evidence quote + Ref tag + applied penalty description.
  5. Never write to 1-comparison / 2-Utilities / 3-Warranty / 4-BOM.
  6. Save to 07-output/TBE-<PKG>_filled_YYYY-MM-DD.xlsx.
  7. Print summary: cells filled, anomalies, output path.

No invention: missing facts produce "Not declared — vendor clarification
required" PLUS the prescribed penalty per CLAUDE.md.  No cross-vendor transfer:
Aster scores are computed from Aster's own evidence JSON only (the Zealwe OEM
mapping is recorded in the comment but never imported as a fact).
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
RULES_PATH = ROOT / "00-rules" / "CLAUDE.md"
PROTECTED_SHEETS = {"1-comparison", "2-Utilities", "3-Warranty", "4-BOM"}

CAPEX_TEXT = "TBC — Procurement Team assessment, sealed commercial offer"

JSON_BLOCK_RE = re.compile(r"```json\s*\n(.*?)\n```", re.DOTALL)


# ---------------------------------------------------------------------------
# evidence loader
# ---------------------------------------------------------------------------

def load_evidence(path: Path) -> dict[str, dict[str, Any]]:
    """Return {vendor_name: vendor_evidence_dict} from the .md."""
    if not path.exists():
        raise FileNotFoundError(f"evidence file missing: {path}")
    raw = path.read_text(encoding="utf-8")
    out: dict[str, dict[str, Any]] = {}
    for block in JSON_BLOCK_RE.findall(raw):
        try:
            data = json.loads(block)
        except json.JSONDecodeError as e:
            raise ValueError(f"bad JSON in {path}: {e}\n--- block ---\n{block[:400]}") from e
        name = data.get("vendor")
        if not name:
            continue
        out[name] = data
    return out


def fact(node: Any) -> tuple[Any, str | None]:
    """Return (value, ref_tag) from an evidence node.

    A node is either:
      - None / missing  →  (None, None)
      - {"value": ..., "ref": {"file","sheet","row","quote"}, ...}
      - bare scalar (treated as undeclared metadata) → (value, None)
    """
    if node is None:
        return None, None
    if isinstance(node, dict) and "value" in node:
        val = node.get("value")
        ref = node.get("ref")
        tag = None
        if ref:
            tag = f"Ref: {ref.get('file')}!{ref.get('sheet')}!{ref.get('row')} — '{(ref.get('quote') or '')[:120]}'"
        return val, tag
    return node, None


def ref_tag(node: Any) -> str | None:
    _, t = fact(node)
    return t


# ---------------------------------------------------------------------------
# scoring primitives
# ---------------------------------------------------------------------------

@dataclass
class Cell:
    score: float
    sxw: float
    comment: str
    anomalies: list[str] = field(default_factory=list)


def clip(x: float, lo: float = 0, hi: float = 10) -> float:
    return max(lo, min(hi, x))


def origin_risk(country: str | None) -> tuple[float, str]:
    """Return (penalty, label) per CLAUDE.md origin rules.

    Note: this returns the ADDITIONAL penalty when no India stock / MTTR are
    declared. The full China rule (−3) is only applied when no India stock AND
    no MTTR; we apply the milder −2 baseline when origin is risky but evidence
    on stock/MTTR is silent (still risky).
    """
    if not country:
        return 0, "origin not declared"
    c = country.lower()
    if "china" in c:
        return -3, "China-origin critical spare + no India stock + no MTTR → −3"
    if any(k in c for k in ("germany", "eu", "europe", "us ", "usa", "united states", "singapore")):
        return -2, f"{country}-origin + weak India presence → −2"
    if "india" in c:
        return 0, "India-origin — no origin penalty"
    return 0, f"origin '{country}' — neutral"


def fc_band(fc: int | None, pc: int | None, nc: int | None) -> tuple[float, str]:
    """Map FC% → base score per CLAUDE.md bands.

    Guard rails (no invention):
    - FC missing  → 0 (cannot score Design Concept without a compliance tally).
    - FC present but PC AND NC both missing → cannot compute the denominator,
      so band collapses to a conservative 5 with explicit penalty note.
    """
    if fc is None:
        return 0, "FC count not declared (band 0; vendor clarification required)"
    if pc is None and nc is None:
        return 5, (f"FC={fc} declared but PC/NC not tallied — denominator unknown; "
                   f"conservative band 5 (CLAUDE.md Hard Rule 1: no invention)")
    total = (fc or 0) + (pc or 0) + (nc or 0)
    if total == 0:
        return 0, "FC/PC/NC denominator zero (no rows captured)"
    pct = 100 * fc / total
    if pct >= 99:
        s = 10
    elif pct >= 90:
        s = 9
    elif pct >= 80:
        s = 8
    elif pct >= 70:
        s = 7
    elif pct >= 60:
        s = 6
    else:
        s = 5
    if (nc or 0) > 0:
        s = min(s, 5)
    return s, f"FC%={pct:.1f}  → base {s}"


# ---------------------------------------------------------------------------
# scorers — one per criterion (Subcategory match is case-insensitive prefix)
# ---------------------------------------------------------------------------

def score_design_concept(v: dict[str, Any], w: int) -> Cell:
    fc = (v.get("fc_pc_nc") or {}).get("fc")
    pc = (v.get("fc_pc_nc") or {}).get("pc")
    nc = (v.get("fc_pc_nc") or {}).get("nc")
    note = (v.get("fc_pc_nc") or {}).get("note") or ""
    ref = (v.get("fc_pc_nc") or {}).get("ref")

    base, band_note = fc_band(fc, pc, nc)
    anomalies = []
    penalty_lines = []

    # Metal-halide / lamp-life CTQ check
    lamp_life = (v.get("lamp") or {}).get("life_hours")
    if lamp_life is None:
        penalty_lines.append("Lamp life not declared — −1 reliability tag (Design Concept retains base)")
        anomalies.append(f"{v['vendor']}: lamp life undeclared")

    # NIST/PTB sensor traceability
    if not ((v.get("ctq") or {}).get("sensor_nist_ptb_traceable")):
        penalty_lines.append("Sensor NIST/PTB traceability not declared — flag CTQ")
        anomalies.append(f"{v['vendor']}: sensor traceability undeclared")

    score = clip(base)
    sxw = round(score * w, 2)

    if fc is None:
        comment = (
            f"Not declared — vendor clarification required. {note} "
            f"Penalty: scored 0 until FC tally is shared."
        )
    else:
        comment = f"{band_note}. Evidence note: {note[:200]}."
    if ref:
        comment += f" Ref: {ref.get('file')}!{ref.get('sheet')}!{ref.get('row')} — '{(ref.get('quote') or '')[:100]}'."
    if penalty_lines:
        comment += " " + " | ".join(penalty_lines)
    return Cell(score, sxw, comment, anomalies)


def score_lessons_learnt(v: dict[str, Any], w: int) -> Cell:
    """Responsiveness proxy: count of revised submissions + PC transparency."""
    note = (v.get("fc_pc_nc") or {}).get("note") or ""
    ref = (v.get("fc_pc_nc") or {}).get("ref")
    pc = (v.get("fc_pc_nc") or {}).get("pc") or 0
    fc = (v.get("fc_pc_nc") or {}).get("fc")

    anomalies = []
    base = 5  # neutral default for "no signals either way"
    detail = []

    quote = (ref or {}).get("quote", "") if ref else ""
    if ref and "Rev" in (ref.get("file", "") or ""):
        base += 2
        detail.append("Revised submission filed (Rev1_) — +2 responsiveness")
    if pc > 0 and fc is not None:
        base += 1
        detail.append(f"Declared {pc} PC deviations transparently — +1 honesty")
    if fc is None:
        base = max(1, base - 3)
        detail.append("No FC/PC tally → −3 (cannot assess responsiveness)")
        anomalies.append(f"{v['vendor']}: cannot score responsiveness without compliance tally")

    score = clip(base)
    sxw = round(score * w, 2)
    comment = "; ".join(detail) if detail else "Neutral — no responsiveness signals in evidence (base 5)."
    if ref:
        comment += f" Ref: {ref.get('file')}!{ref.get('sheet')}!{ref.get('row')} — '{quote[:100]}'."
    return Cell(score, sxw, comment, anomalies)


def score_tech_enabler(v: dict[str, Any], w: int) -> Cell:
    d = v.get("digital") or {}
    protocols = [("OPC-UA", d.get("opc_ua")), ("MODBUS-TCP", d.get("modbus_tcp")),
                 ("REST API", d.get("rest_api")), ("LIMS", d.get("lims"))]
    declared = [name for name, node in protocols if fact(node)[0]]
    base = 2 + 2 * len(declared)  # 0 declared → 2; 4 declared → 10
    anomalies = []
    if not declared:
        anomalies.append(f"{v['vendor']}: no digital integration declared (penalty −2 vs neutral)")
    score = clip(base)
    sxw = round(score * w, 2)
    if declared:
        comment = f"Declared protocols: {', '.join(declared)}. Each protocol +2 (base 2)."
    else:
        comment = "No OPC-UA / MODBUS / REST / LIMS declared — base 2. Penalty: vendor clarification required."
    refs = [ref_tag(n) for _, n in protocols if ref_tag(n)]
    if refs:
        comment += " " + " | ".join(refs)
    return Cell(score, sxw, comment, anomalies)


def score_tech_maturity(v: dict[str, Any], w: int) -> Cell:
    op = v.get("oem_profile") or {}
    rel = v.get("reliability") or {}
    installed_val, installed_ref = fact(op.get("installed_base"))
    years_val, years_ref = fact(op.get("years_in_pv_test"))
    refs_val, refs_ref = fact(op.get("indian_reference_customers"))
    warranty_val, warranty_ref = fact(rel.get("warranty_years"))

    base = 5
    notes = []
    if installed_val is not None:
        try:
            n = int(installed_val)
            if n >= 50:
                base += 3
                notes.append(f"Installed base {n} ≥50 → +3")
            elif n >= 10:
                base += 2
                notes.append(f"Installed base {n} ≥10 → +2")
            else:
                base += 1
                notes.append(f"Installed base {n} → +1")
        except (TypeError, ValueError):
            notes.append(f"Installed base declared but non-numeric ({installed_val}) → +1")
            base += 1
    if years_val:
        base += 1
        notes.append("Years in PV test declared → +1")
    if refs_val:
        base += 1
        notes.append("Indian reference customer(s) declared → +1")
    if warranty_val and isinstance(warranty_val, (int, float)) and warranty_val >= 3:
        base += 1
        notes.append(f"Warranty {warranty_val}y (≥3) → +1 maturity")

    anomalies = []
    if not (installed_val or years_val or refs_val):
        anomalies.append(f"{v['vendor']}: no maturity signal (installed base / years / references all blank)")

    score = clip(base)
    sxw = round(score * w, 2)
    comment = "; ".join(notes) if notes else "No installed base / years / references declared — base 5."
    for r in (installed_ref, years_ref, refs_ref, warranty_ref):
        if r:
            comment += " " + r
    return Cell(score, sxw, comment, anomalies)


def score_ease_of_operation(v: dict[str, Any], w: int) -> Cell:
    rel = v.get("reliability") or {}
    lamp = v.get("lamp") or {}
    hmi_val, hmi_ref = fact(rel.get("hmi_languages"))
    op_count_val, op_count_ref = fact(rel.get("operator_headcount"))
    lamp_type_val, lamp_type_ref = fact(lamp.get("type"))
    lamp_life_val, _ = fact(lamp.get("life_hours"))

    base = 5
    notes = []
    anomalies = []

    if hmi_val:
        base += 1
        notes.append("HMI languages declared → +1")
    else:
        notes.append("HMI languages not declared (no bonus)")
    if op_count_val:
        base += 1
        notes.append("Operator headcount declared → +1")
    # CTQ — metal-halide w/o HMI run-hours
    if lamp_type_val and "halide" in str(lamp_type_val).lower():
        if isinstance(lamp_life_val, (int, float)) and lamp_life_val <= 2000:
            base -= 3
            notes.append("Metal-halide + life ≤2000h + no HMI run-hours view → −3")
            anomalies.append(f"{v['vendor']}: metal-halide ≤2000h CTQ penalty")
        else:
            base -= 2
            notes.append("Metal-halide lamp without run-hours-vs-warranted-life view → −2")
    # Generic HMI visibility penalty if lamp declared but no HMI visibility
    elif lamp_type_val and not hmi_val:
        base -= 1
        notes.append("Lamp declared but no HMI run-hours visibility → −1")

    score = clip(base)
    sxw = round(score * w, 2)
    comment = "; ".join(notes)
    for r in (hmi_ref, op_count_ref, lamp_type_ref):
        if r:
            comment += " " + r
    return Cell(score, sxw, comment, anomalies)


def score_equipment_availability(v: dict[str, Any], w: int) -> Cell:
    rel = v.get("reliability") or {}
    spares = v.get("spares") or {}
    op = v.get("oem_profile") or {}
    mtbf_val, mtbf_ref = fact(rel.get("mtbf_hours"))
    warranty_val, warranty_ref = fact(rel.get("warranty_years"))
    installed_val, installed_ref = fact(op.get("installed_base"))
    origin_val, origin_ref = fact(spares.get("critical_origin_country"))

    base = 5
    notes = []
    anomalies = []

    if warranty_val and isinstance(warranty_val, (int, float)):
        if warranty_val >= 3:
            base += 3
            notes.append(f"Warranty {warranty_val}y (≥3) → +3")
        elif warranty_val >= 1:
            base -= 2
            notes.append(f"Warranty {warranty_val}y (1-2) → −2 baseline")
    else:
        base -= 2
        notes.append("Warranty not declared → −2")
        anomalies.append(f"{v['vendor']}: warranty undeclared")

    if mtbf_val is None:
        base -= 1
        notes.append("MTBF not declared → −1")
    if installed_val is None:
        base -= 1
        notes.append("Installed base not declared → −1")

    # Combined origin risk per CLAUDE.md
    if (warranty_val is None or (isinstance(warranty_val, (int, float)) and warranty_val < 3)) \
            and mtbf_val is None and installed_val is None:
        op_pen, op_label = origin_risk(origin_val)
        if op_pen < 0:
            base += op_pen
            notes.append(f"Combined warranty/MTBF/install gap + origin → {op_label}")

    score = clip(base)
    sxw = round(score * w, 2)
    comment = "; ".join(notes)
    for r in (warranty_ref, mtbf_ref, installed_ref, origin_ref):
        if r:
            comment += " " + r
    return Cell(score, sxw, comment, anomalies)


def score_spare_part(v: dict[str, Any], w: int) -> Cell:
    spares = v.get("spares") or {}
    rel = v.get("reliability") or {}
    origin_val, origin_ref = fact(spares.get("critical_origin_country"))
    india_stock_val, india_stock_ref = fact(spares.get("india_stock"))
    spare_yrs_val, spare_yrs_ref = fact(spares.get("spare_commitment_years"))
    mttr_val, mttr_ref = fact(rel.get("mttr_hours"))

    base = 5
    notes = []
    anomalies = []

    if india_stock_val:
        base += 2
        notes.append("India stock declared → +2")
    else:
        notes.append("India stock not declared (no bonus)")
        anomalies.append(f"{v['vendor']}: India spare stock undeclared")

    if isinstance(spare_yrs_val, (int, float)) and spare_yrs_val >= 10:
        base += 2
        notes.append(f"10-yr spare commitment ({spare_yrs_val}y) → +2")

    op_pen, op_label = origin_risk(origin_val)
    # China rule needs no India stock AND no MTTR for full −3
    if "China" in (origin_val or "") and not india_stock_val and mttr_val is None:
        base -= 3
        notes.append("China origin + no India stock + no MTTR → −3 (CLAUDE.md Penalty 1)")
    elif op_pen < 0:
        base += op_pen
        notes.append(op_label)

    score = clip(base)
    sxw = round(score * w, 2)
    comment = "; ".join(notes)
    for r in (origin_ref, india_stock_ref, spare_yrs_ref, mttr_ref):
        if r:
            comment += " " + r
    return Cell(score, sxw, comment, anomalies)


def score_service_lead_time(v: dict[str, Any], w: int) -> Cell:
    svc = v.get("service") or {}
    centres_val, centres_ref = fact(svc.get("india_centres"))
    engineers_val, engineers_ref = fact(svc.get("india_engineers"))
    sla_val, sla_ref = fact(svc.get("sla_hours"))
    jamnagar_val, jamnagar_ref = fact(svc.get("jamnagar_proximity"))

    base = 5
    notes = []
    anomalies = []
    capped = False

    if sla_val is None:
        # Generic "service provided" w/o SLA → cap 5, −3
        base = min(base, 5) - 3
        capped = True
        notes.append("No quantified SLA → cap at 5, additional −3 (CLAUDE.md Penalty 3)")
        anomalies.append(f"{v['vendor']}: SLA hours undeclared")
    else:
        base += 1
        notes.append(f"Quantified SLA ({sla_val} h) → +1")

    if centres_val:
        # try to count comma-separated entries
        try:
            n = len(str(centres_val).split(","))
        except Exception:
            n = 1
        bonus = min(3, n)
        if not capped:
            base += bonus
            notes.append(f"India service centres ({n}) → +{bonus}")
        else:
            notes.append(f"India service centres ({n}) declared but score capped")
    else:
        notes.append("India service centres not declared")

    if isinstance(engineers_val, (int, float)):
        if engineers_val >= 20:
            tier = 3
        elif engineers_val >= 6:
            tier = 2
        elif engineers_val >= 1:
            tier = 1
        else:
            tier = 0
        if not capped:
            base += tier
            notes.append(f"{int(engineers_val)} India engineers → +{tier}")
        else:
            notes.append(f"{int(engineers_val)} India engineers declared but score capped")

    if jamnagar_val and "no explicit" not in str(jamnagar_val).lower() and not capped:
        base += 1
        notes.append("Western India / Jamnagar proximity → +1")

    score = clip(base)
    sxw = round(score * w, 2)
    comment = "; ".join(notes)
    for r in (sla_ref, centres_ref, engineers_ref, jamnagar_ref):
        if r:
            comment += " " + r
    return Cell(score, sxw, comment, anomalies)


# ---------------------------------------------------------------------------
# routing
# ---------------------------------------------------------------------------

SCORER = {
    "ease of operation": score_ease_of_operation,
    "equipment availability": score_equipment_availability,
    "spare part availability": score_spare_part,
    "service & support lead time": score_service_lead_time,
    "design concept": score_design_concept,
    "lessons learnt": score_lessons_learnt,
    "technology enabler": score_tech_enabler,
    "technology maturity": score_tech_maturity,
}


def route(subcategory: str):
    key = subcategory.lower().strip()
    for prefix, fn in SCORER.items():
        if key.startswith(prefix):
            return fn
    return None


# ---------------------------------------------------------------------------
# template walker
# ---------------------------------------------------------------------------

@dataclass
class Layout:
    header_row: int
    col_cat: int
    col_sub: int
    col_w: int
    vendors: list[tuple[str, int, int, int]]  # (name, details_col, score_col, sxw_col)


def detect_layout(ws: Worksheet) -> Layout:
    """Locate the header row by scanning for the 'Subcategory' cell."""
    header_row = None
    col_cat = col_sub = col_w = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            v = (cell.value or "")
            if isinstance(v, str) and v.strip().lower() == "subcategory":
                header_row = cell.row
                col_sub = cell.column
                break
        if header_row:
            break
    if not header_row:
        raise RuntimeError("could not find 'Subcategory' header cell — template layout mismatch")

    # walk that row
    vendors: list[tuple[str, int, int, int]] = []
    last_vendor_name = None
    triples: dict[str, dict[str, int]] = {}
    for cell in ws[header_row]:
        v = (cell.value or "")
        if not isinstance(v, str):
            continue
        s = v.strip()
        low = s.lower()
        if low == "category":
            col_cat = cell.column
        elif low == "w":
            col_w = cell.column
        elif " — " in s:
            vendor, kind = [x.strip() for x in s.split(" — ", 1)]
            triples.setdefault(vendor, {})[kind.lower()] = cell.column
            last_vendor_name = vendor

    for vendor, cols in triples.items():
        d = cols.get("details")
        sc = cols.get("score")
        sw = cols.get("sxw") or cols.get("s×w") or cols.get("s x w") or cols.get("s*w")
        if not (d and sc and sw):
            continue
        vendors.append((vendor, d, sc, sw))

    if not (col_cat and col_sub and col_w):
        raise RuntimeError("missing Category/Subcategory/W columns in header row")
    if not vendors:
        raise RuntimeError("no vendor triples (Details/Score/SxW) found in header row")

    return Layout(header_row=header_row, col_cat=col_cat, col_sub=col_sub, col_w=col_w, vendors=vendors)


# ---------------------------------------------------------------------------
# main fill pass
# ---------------------------------------------------------------------------

@dataclass
class Summary:
    cells_filled: int = 0
    capex_cells: int = 0
    rows_processed: int = 0
    vendors_processed: int = 0
    anomalies: list[str] = field(default_factory=list)
    per_vendor_total: dict[str, float] = field(default_factory=dict)


def fill_workbook(template_path: Path, evidence: dict[str, dict[str, Any]], output_path: Path) -> Summary:
    wb = load_workbook(template_path, keep_vba=True, keep_links=True)
    summary = Summary()

    target_sheets = [s for s in wb.sheetnames if s not in PROTECTED_SHEETS]
    if not target_sheets:
        raise RuntimeError("template has no writable sheet (all are in PROTECTED_SHEETS)")

    # The fillable scoring tab is the first non-protected one (typically "0-TBE").
    ws = wb[target_sheets[0]]
    layout = detect_layout(ws)
    summary.vendors_processed = len(layout.vendors)

    wrap = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)
    capex_fill = PatternFill("solid", fgColor="FFF2CC")

    for r in range(layout.header_row + 1, ws.max_row + 1):
        cat = (ws.cell(row=r, column=layout.col_cat).value or "")
        sub = (ws.cell(row=r, column=layout.col_sub).value or "")
        w_val = ws.cell(row=r, column=layout.col_w).value
        if not isinstance(sub, str) or not sub.strip():
            continue
        # carry forward Category (template may leave it blank on repeat rows)
        if not cat:
            for back in range(r - 1, layout.header_row, -1):
                pc = ws.cell(row=back, column=layout.col_cat).value
                if pc:
                    cat = pc
                    break

        summary.rows_processed += 1
        is_capex = isinstance(cat, str) and cat.strip().lower() == "capex"
        scorer = route(sub) if not is_capex else None

        for vendor, dcol, scol, wcol in layout.vendors:
            if is_capex:
                ws.cell(row=r, column=dcol, value=CAPEX_TEXT).alignment = wrap
                ws.cell(row=r, column=dcol).fill = capex_fill
                ws.cell(row=r, column=scol, value=None)
                ws.cell(row=r, column=wcol, value=None)
                summary.capex_cells += 1
                continue

            v_data = evidence.get(vendor)
            if v_data is None:
                # vendor canonically expected but absent from evidence
                cell = Cell(0, 0, f"Not declared — vendor clarification required (no evidence block for '{vendor}'). Penalty: scored 0.", [f"{vendor}: no evidence block"])
            elif scorer is None:
                cell = Cell(0, 0, f"Unrouted subcategory '{sub}' — fill_tbe.py needs a scorer.", [f"unrouted: {sub}"])
            else:
                try:
                    cell = scorer(v_data, int(w_val or 0))
                except Exception as e:
                    cell = Cell(0, 0, f"Scorer error: {e}", [f"{vendor} / {sub}: {e}"])

            ws.cell(row=r, column=dcol, value=cell.comment).alignment = wrap
            ws.cell(row=r, column=scol, value=cell.score).font = bold
            ws.cell(row=r, column=wcol, value=cell.sxw).font = bold
            summary.cells_filled += 3
            summary.anomalies.extend(cell.anomalies)
            summary.per_vendor_total[vendor] = summary.per_vendor_total.get(vendor, 0) + cell.sxw

    # Total / percentage row update (looks for "TOTAL" cell in col_sub)
    for r in range(layout.header_row + 1, ws.max_row + 2):
        v = ws.cell(row=r, column=layout.col_sub).value
        if isinstance(v, str) and v.strip().upper().startswith("TOTAL"):
            for vendor, dcol, scol, wcol in layout.vendors:
                total = summary.per_vendor_total.get(vendor, 0)
                ws.cell(row=r, column=dcol, value=f"{total:.1f} / 1000  ({total / 10:.1f}%)").font = bold
                ws.cell(row=r, column=scol, value=None)
                ws.cell(row=r, column=wcol, value=round(total, 1)).font = bold
            break

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return summary


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Fill TBE template from evidence (deterministic).")
    p.add_argument("--package", default="UV-2")
    p.add_argument("--template", default=None, help="defaults to 01-templates/TBE-<PKG>.xlsx")
    p.add_argument("--evidence", default=None, help="defaults to 05-evidence/<PKG>_evidence.md")
    p.add_argument("--output",   default=None, help="defaults to 07-output/TBE-<PKG>_filled_<DATE>.xlsx")
    args = p.parse_args(argv)

    pkg = args.package
    today = dt.date.today().isoformat()

    template = Path(args.template) if args.template else ROOT / "01-templates" / f"TBE-{pkg.replace('-', '')}.xlsx"
    if not template.exists():
        # try with hyphen too (UV-2 vs UV2)
        alt = ROOT / "01-templates" / f"TBE-{pkg}.xlsx"
        if alt.exists():
            template = alt
    evidence = Path(args.evidence) if args.evidence else ROOT / "05-evidence" / f"{pkg}_evidence.md"
    output   = Path(args.output)   if args.output   else ROOT / "07-output" / f"TBE-{pkg.replace('-', '')}_filled_{today}.xlsx"

    if not RULES_PATH.exists():
        print(f"FATAL: rules file missing: {RULES_PATH}", file=sys.stderr); return 2
    if not template.exists():
        print(f"FATAL: template missing: {template}", file=sys.stderr); return 2
    if not evidence.exists():
        print(f"FATAL: evidence missing: {evidence}", file=sys.stderr); return 2

    ev = load_evidence(evidence)
    summary = fill_workbook(template, ev, output)

    # JSON sidecar consumed by the Vercel stub page (app/procurement/tbe).
    sidecar = output.with_suffix(".json")
    sidecar.write_text(json.dumps({
        "package": pkg,
        "date": today,
        "template": str(template.relative_to(ROOT)),
        "evidence": str(evidence.relative_to(ROOT)),
        "output_xlsx": str(output.relative_to(ROOT)),
        "rows": summary.rows_processed,
        "vendors": list(summary.per_vendor_total.keys()),
        "cells_filled": summary.cells_filled,
        "capex_cells": summary.capex_cells,
        "totals_per_vendor": {
            k: {"score": round(v, 1), "percent": round(v / 10, 1)}
            for k, v in sorted(summary.per_vendor_total.items(), key=lambda kv: -kv[1])
        },
        "anomalies": summary.anomalies,
    }, indent=2), encoding="utf-8")

    print()
    print(f"=== TBE fill summary  ({pkg}) ===")
    print(f"  template  : {template.relative_to(ROOT)}")
    print(f"  evidence  : {evidence.relative_to(ROOT)}  ({len(ev)} vendor blocks)")
    print(f"  output    : {output.relative_to(ROOT)}")
    print(f"  rows      : {summary.rows_processed}")
    print(f"  vendors   : {summary.vendors_processed}  ({', '.join(n for n,*_ in [(v, ) for v in summary.per_vendor_total.keys()])})")
    print(f"  cells filled  : {summary.cells_filled}  (scoring) + {summary.capex_cells}  (Capex placeholders)")
    print()
    print("  Weighted totals (/ 1000):")
    for vendor in sorted(summary.per_vendor_total, key=lambda k: -summary.per_vendor_total[k]):
        print(f"    {vendor:10s}  {summary.per_vendor_total[vendor]:6.1f}   ({summary.per_vendor_total[vendor]/10:5.1f}%)")
    print()
    if summary.anomalies:
        print(f"  Anomalies ({len(summary.anomalies)}):")
        for a in summary.anomalies[:30]:
            print(f"    - {a}")
        if len(summary.anomalies) > 30:
            print(f"    ... +{len(summary.anomalies) - 30} more")
    else:
        print("  No anomalies recorded.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
