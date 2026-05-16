"""TBE Excel writer.

For each package this module either:
  * loads a master TBE template from ``templates/<package>.xlsm`` (or
    ``.xlsx``) -- preserving VBA, links, formulas, merged cells, and
    embedded images -- and *modifies* it in place; or
  * if no master template is available, materialises a minimal but
    well-formed TBE workbook from scratch using the canonical row order
    declared in this module.

The writer never regenerates a template from scratch when one exists.
That guarantee is enforced in :func:`load_or_build_workbook` and tested
in ``tests/test_round_trip.py``.

The output is one workbook per package:
``deliverables/TBE-<PKG>-v<YYYYMMDD>.xlsx`` plus a Master Summary
workbook ``deliverables/TBE-Master-Summary-v<YYYYMMDD>.xlsx``.
"""

from __future__ import annotations

import argparse
import json
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.workbook import Workbook as WbType

from .packages import (
    CAPEX_PLACEHOLDER,
    NOT_DECLARED,
    Package,
    PACKAGES,
    by_key,
)
from .scoring import (
    ScoreLine,
    capex_row,
    compliance_tally_score,
    fluorescent_uv_adjustment,
    make_ref,
    multilingual_hmi_credit,
    option_a_penalty,
    reliability_score,
)


log = logging.getLogger("fill_tbe")

TEMPLATES_ROOT = Path("templates")
DELIVERABLES_ROOT = Path("deliverables")
EVIDENCE_ROOT = Path("evidence")


# ---------------------------------------------------------------------------
# Row plan
# ---------------------------------------------------------------------------


# Canonical TBE row order. Rows added or removed here propagate to every
# package; section names cluster rows in the output.
ROW_PLAN: tuple[tuple[str, str], ...] = (
    ("Design Concept", "Compliance tally (FC/PC/DNC)"),
    ("Design Concept", "Thermal design (refrigerant/compressor/MoC/uniformity)"),
    ("Design Concept", "Connectivity (OPC-UA / MODBUS / REST / LIMS)"),
    ("Design Concept", "Installed base (India + Global)"),
    ("Design Concept", "Spectral match (SSS-only)"),
    ("Reliability",    "MTBF (hours)"),
    ("Reliability",    "Warranty (months)"),
    ("Reliability",    "HMI run-hour visibility"),
    ("Reliability",    "Service network (centres / SLA)"),
    ("Risk",           "Option-A (BoM origin)"),
    ("Risk",           "CTQ deviations"),
    ("Ease of Operation", "Multilingual HMI credit"),
    ("Ease of Operation", "Lamp / LED life (SSS-only)"),
    ("Capex",          "Equipment purchase price"),
    ("Capex",          "Installation & commissioning"),
    ("Capex",          "Annual maintenance contract (AMC)"),
)


HEADER_ROW = (
    "Section",
    "Criterion",
    "Quoted value",
    "Score",
    "Ref",
    "Comment",
)


# ---------------------------------------------------------------------------
# Template loading
# ---------------------------------------------------------------------------


def _template_path(package: Package) -> Path | None:
    for ext in (".xlsm", ".xlsx"):
        p = TEMPLATES_ROOT / f"{package.key}{ext}"
        if p.exists():
            return p
    return None


def load_or_build_workbook(package: Package) -> tuple[WbType, bool]:
    """Return ``(workbook, from_template)``.

    If a master template exists it is loaded with ``keep_vba=True`` and
    ``keep_links=True`` so VBA macros, external links, formulas, merged
    cells and embedded images survive intact. Otherwise a minimal
    workbook is built.
    """
    path = _template_path(package)
    if path is not None:
        is_macro = path.suffix.lower() == ".xlsm"
        wb = load_workbook(
            filename=path,
            keep_vba=is_macro,
            keep_links=True,
            data_only=False,
        )
        return wb, True
    wb = Workbook()
    return wb, False


# ---------------------------------------------------------------------------
# Score computation per vendor
# ---------------------------------------------------------------------------


@dataclass
class VendorScore:
    vendor: str
    lines: list[ScoreLine]
    design_total: float
    reliability_total: float
    ease_total: float
    risk_total: float
    capex_total: float
    grand_total: float
    coverage_pct: float


def _read_evidence(package: Package, vendor: str, evidence_root: Path) -> dict[str, Any]:
    path = evidence_root / package.key / f"{_safe(vendor)}.json"
    if not path.exists():
        log.warning("No evidence file at %s; treating all fields as Not declared.", path)
        return {}
    return json.loads(path.read_text())


def _field_has_evidence(field: dict[str, Any] | None) -> bool:
    """Return True if the field actually carries vendor evidence.

    A field that came from the 'no-pinecone' or 'error:*' fallback paths
    has every schema key set to ``NOT_DECLARED``, including
    ``source_file``. Treat such fields the same as missing -- the row
    must read ``Not declared`` and accrue zero contribution.
    """
    if not field or not isinstance(field, dict):
        return False
    value = field.get("value")
    if not isinstance(value, dict) or not value:
        return False
    # If source_file is Not declared the rest of the value is too.
    if value.get("source_file", NOT_DECLARED) == NOT_DECLARED:
        return False
    return True


def _safe(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in name)


def score_vendor(package: Package, vendor: str, evidence: dict[str, Any]) -> VendorScore:
    tally = evidence.get("compliance_tally")
    thermal = evidence.get("thermal_design")
    reliability = evidence.get("reliability")
    bom = evidence.get("bom_origin")
    service = evidence.get("service_network")
    conn = evidence.get("connectivity")
    installed = evidence.get("installed_base")
    deviations = evidence.get("ctq_deviations")
    spectral = evidence.get("spectral_match")
    lamp_life = evidence.get("lamp_life")

    lines: list[ScoreLine] = []

    # --- Design Concept --------------------------------------------------
    tally_line = compliance_tally_score(tally)
    # Rule 3: fluorescent-UV downgrade folds into the compliance tally
    # row (which represents Design Concept here).
    adjusted, comment = fluorescent_uv_adjustment(package, tally_line.score, thermal)
    if comment:
        tally_line = ScoreLine(tally_line.criterion, tally_line.value, adjusted,
                                tally_line.ref, (tally_line.comment + " " + comment).strip())
    lines.append(tally_line)

    # Thermal design: declared/not-declared row, not numeric.
    if _field_has_evidence(thermal):
        v = thermal.get("value", {})
        thermal_value = (
            f"refrigerant={v.get('refrigerant', NOT_DECLARED)}; "
            f"compressor/source={v.get('compressor', NOT_DECLARED)}; "
            f"MoC={v.get('moc', NOT_DECLARED)}; "
            f"uniformity={v.get('uniformity', NOT_DECLARED)}"
        )
        thermal_score = 0.0 if NOT_DECLARED in thermal_value else 2.0
        lines.append(ScoreLine(
            "Thermal design (refrigerant/compressor/MoC/uniformity)",
            thermal_value, thermal_score, make_ref(thermal),
        ))
    else:
        lines.append(ScoreLine(
            "Thermal design (refrigerant/compressor/MoC/uniformity)",
            NOT_DECLARED, 0.0, make_ref(thermal), "No thermal evidence; not declared.",
        ))

    # Connectivity
    if _field_has_evidence(conn):
        v = conn.get("value", {})
        flags = [k for k in ("opc_ua", "modbus_rtu", "modbus_tcp", "rest_api", "lims") if v.get(k) is True]
        score = float(len(flags))  # 0-5
        lines.append(ScoreLine("Connectivity (OPC-UA / MODBUS / REST / LIMS)",
                                ", ".join(flags) or "None declared",
                                score, make_ref(conn)))
    else:
        lines.append(ScoreLine("Connectivity (OPC-UA / MODBUS / REST / LIMS)",
                                NOT_DECLARED, 0.0, make_ref(conn), "Not declared."))

    # Installed base
    if _field_has_evidence(installed):
        v = installed.get("value", {})
        try:
            india = int(v.get("india_installs", 0) or 0)
            world = int(v.get("global_installs", 0) or 0)
        except (TypeError, ValueError):
            india = world = 0
        score = min(3.0, india * 0.5 + world * 0.05)
        lines.append(ScoreLine("Installed base (India + Global)",
                                f"India={india}, Global={world}",
                                round(score, 2), make_ref(installed)))
    else:
        lines.append(ScoreLine("Installed base (India + Global)",
                                NOT_DECLARED, 0.0, make_ref(installed), "Not declared."))

    # Spectral match -- only meaningful for sun-simulators.
    if package.is_sun_simulator:
        if _field_has_evidence(spectral):
            v = spectral.get("value", {})
            cls = str(v.get("spectral_class", NOT_DECLARED)).upper()
            # AAA = 3 pts; AAB = 2; ABB = 1; BBB = 0; CCC = -1.
            mapping = {"AAA": 3.0, "AAB": 2.0, "ABA": 2.0, "BAA": 2.0,
                       "ABB": 1.0, "BAB": 1.0, "BBA": 1.0, "BBB": 0.0,
                       "CCC": -1.0}
            score = mapping.get(cls, 0.0)
            lines.append(ScoreLine("Spectral match (SSS-only)", cls, score, make_ref(spectral)))
        else:
            lines.append(ScoreLine("Spectral match (SSS-only)", NOT_DECLARED, 0.0,
                                    make_ref(spectral), "Not declared."))
    else:
        lines.append(ScoreLine("Spectral match (SSS-only)", "n/a (not a sun simulator)", 0.0,
                                "Ref: n/a", "Row only scored for SSS packages."))

    # --- Reliability -----------------------------------------------------
    rel_lines = reliability_score(reliability, lamp_life)
    lines.extend(rel_lines)

    # Service network
    if _field_has_evidence(service):
        v = service.get("value", {})
        try:
            centres = int(v.get("centres", 0) or 0)
            sla = float(v.get("sla_hours", 9999) or 9999)
        except (TypeError, ValueError):
            centres = 0
            sla = 9999
        score = 0.0
        if centres > 0:
            score += min(2.0, centres * 0.5)
        if sla <= 24:
            score += 1.0
        if v.get("support_24x7") is True:
            score += 1.0
        lines.append(ScoreLine(
            "Service network (centres / SLA)",
            f"centres={centres}, SLA={v.get('sla_hours', NOT_DECLARED)}h, "
            f"24x7={v.get('support_24x7', NOT_DECLARED)}",
            round(score, 2), make_ref(service),
        ))
    else:
        lines.append(ScoreLine("Service network (centres / SLA)", NOT_DECLARED, 0.0,
                                make_ref(service), "Not declared."))

    # --- Risk ------------------------------------------------------------
    lines.append(option_a_penalty(bom))
    if _field_has_evidence(deviations):
        v = deviations.get("value", {})
        devs = v.get("deviations", [])
        if isinstance(devs, list):
            score = -1.0 * len(devs)
            preview = "; ".join(
                f"{d.get('ctq', '?')}: {d.get('rfq_value', '?')}->{d.get('offered_value', '?')}"
                for d in devs[:3] if isinstance(d, dict)
            )
            lines.append(ScoreLine("CTQ deviations", preview or "None", score, make_ref(deviations)))
        else:
            lines.append(ScoreLine("CTQ deviations", NOT_DECLARED, 0.0, make_ref(deviations),
                                    "Deviations field not a list; treated as not declared."))
    else:
        lines.append(ScoreLine("CTQ deviations", NOT_DECLARED, 0.0, make_ref(deviations),
                                "Not declared."))

    # --- Ease of Operation ----------------------------------------------
    lines.append(multilingual_hmi_credit(vendor, conn))
    if package.is_sun_simulator:
        if _field_has_evidence(lamp_life):
            v = lamp_life.get("value", {})
            try:
                hrs = float(v.get("lamp_life_hours", 0) or 0)
            except (TypeError, ValueError):
                hrs = 0
            score = min(3.0, hrs / 2000.0)
            lines.append(ScoreLine("Lamp / LED life (SSS-only)",
                                    f"{v.get('lamp_life_hours', NOT_DECLARED)} h",
                                    round(score, 2), make_ref(lamp_life)))
        else:
            lines.append(ScoreLine("Lamp / LED life (SSS-only)", NOT_DECLARED, 0.0,
                                    make_ref(lamp_life), "Not declared."))
    else:
        lines.append(ScoreLine("Lamp / LED life (SSS-only)", "n/a (not a sun simulator)", 0.0,
                                "Ref: n/a", "Row only scored for SSS packages."))

    # --- Capex (always blank) -------------------------------------------
    lines.append(capex_row("Equipment purchase price"))
    lines.append(capex_row("Installation & commissioning"))
    lines.append(capex_row("Annual maintenance contract (AMC)"))

    # --- Totals ----------------------------------------------------------
    design_total = sum(l.score for l in lines if _section_of(l.criterion) == "Design Concept")
    reliability_total = sum(l.score for l in lines if _section_of(l.criterion) == "Reliability")
    ease_total = sum(l.score for l in lines if _section_of(l.criterion) == "Ease of Operation")
    risk_total = sum(l.score for l in lines if _section_of(l.criterion) == "Risk")
    capex_total = 0.0  # always 0 -- Capex is never auto-scored

    # Cap Design Concept and Ease of Operation by the package weight profile.
    design_total = min(float(package.weights.design_concept), design_total)
    ease_total = min(float(package.weights.ease_of_operation), ease_total)

    grand_total = design_total + reliability_total + ease_total + risk_total + capex_total

    coverage = _coverage(lines)

    return VendorScore(
        vendor=vendor,
        lines=lines,
        design_total=round(design_total, 2),
        reliability_total=round(reliability_total, 2),
        ease_total=round(ease_total, 2),
        risk_total=round(risk_total, 2),
        capex_total=0.0,
        grand_total=round(grand_total, 2),
        coverage_pct=round(coverage, 1),
    )


def _section_of(criterion: str) -> str:
    for section, label in ROW_PLAN:
        if label == criterion:
            return section
    return "Other"


def _coverage(lines: list[ScoreLine]) -> float:
    """Percent of non-Capex rows with declared (non-'Not declared') values."""
    scored = [l for l in lines if not l.criterion.startswith(("Equipment purchase", "Installation", "Annual maintenance"))]
    if not scored:
        return 0.0
    declared = sum(1 for l in scored if l.value != NOT_DECLARED and l.value != "Not declared")
    return 100.0 * declared / len(scored)


# ---------------------------------------------------------------------------
# Workbook writing
# ---------------------------------------------------------------------------


def write_workbook(package: Package, vendor_scores: list[VendorScore], out_path: Path) -> Path:
    wb, from_template = load_or_build_workbook(package)

    sheet_name = "TBE"
    if from_template:
        # If the template already has a 'TBE' sheet, modify it -- else
        # add a fresh sheet alongside whatever else the template ships
        # (e.g. instructions, RFQ summary, signature block).
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Don't blow away the existing rows; append vendor cols.
            _append_vendor_columns(ws, vendor_scores)
        else:
            ws = wb.create_sheet(sheet_name)
            _write_fresh_tbe_sheet(ws, package, vendor_scores)
    else:
        # Fresh workbook -- drop the default 'Sheet' and create 'TBE'.
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(sheet_name)
        _write_fresh_tbe_sheet(ws, package, vendor_scores)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Wrote TBE workbook: %s (from_template=%s)", out_path, from_template)
    return out_path


def _write_fresh_tbe_sheet(ws, package: Package, vendor_scores: list[VendorScore]) -> None:
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title block
    ws["A1"] = f"Krayavikrayam -- TBE -- {package.display_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"Weights: Design Concept = {package.weights.design_concept}, "
        f"Ease of Operation = {package.weights.ease_of_operation}. "
        f"Capex rows: '{CAPEX_PLACEHOLDER}'."
    )
    ws.merge_cells("A2:F2")

    # Header row at row 4
    header_row = 4
    headers = list(HEADER_ROW)
    for vendor in vendor_scores:
        headers.extend([f"{vendor.vendor} value", f"{vendor.vendor} score",
                        f"{vendor.vendor} ref", f"{vendor.vendor} comment"])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = bold
        cell.fill = fill
        cell.alignment = center

    # Body
    for offset, (section, criterion) in enumerate(ROW_PLAN, start=1):
        row = header_row + offset
        ws.cell(row=row, column=1, value=section)
        ws.cell(row=row, column=2, value=criterion)
        # Columns 3-6 are the per-package "reference" view -- we leave
        # them empty so the per-vendor columns own the actual data.
        ws.cell(row=row, column=5, value="")
        # Vendor columns
        for v_idx, vendor in enumerate(vendor_scores):
            line = next((l for l in vendor.lines if l.criterion == criterion), None)
            base_col = 7 + v_idx * 4
            if line is None:
                ws.cell(row=row, column=base_col, value=NOT_DECLARED)
                ws.cell(row=row, column=base_col + 1, value=0.0)
                ws.cell(row=row, column=base_col + 2, value=f"Ref: {NOT_DECLARED}")
                ws.cell(row=row, column=base_col + 3, value="Row missing from vendor score.")
            else:
                ws.cell(row=row, column=base_col, value=line.value if not isinstance(line.value, (list, dict)) else str(line.value))
                ws.cell(row=row, column=base_col + 1, value=line.score)
                ws.cell(row=row, column=base_col + 2, value=line.ref)
                ws.cell(row=row, column=base_col + 3, value=line.comment)

    # Totals block
    total_row = header_row + len(ROW_PLAN) + 2
    ws.cell(row=total_row, column=2, value="Design Concept total").font = bold
    ws.cell(row=total_row + 1, column=2, value="Reliability total").font = bold
    ws.cell(row=total_row + 2, column=2, value="Ease of Operation total").font = bold
    ws.cell(row=total_row + 3, column=2, value="Risk total").font = bold
    ws.cell(row=total_row + 4, column=2, value="Capex total").font = bold
    ws.cell(row=total_row + 5, column=2, value="GRAND TOTAL").font = Font(bold=True, size=12)
    ws.cell(row=total_row + 6, column=2, value="Evidence coverage %").font = bold

    for v_idx, vendor in enumerate(vendor_scores):
        base_col = 7 + v_idx * 4 + 1  # score column for vendor
        ws.cell(row=total_row, column=base_col, value=vendor.design_total)
        ws.cell(row=total_row + 1, column=base_col, value=vendor.reliability_total)
        ws.cell(row=total_row + 2, column=base_col, value=vendor.ease_total)
        ws.cell(row=total_row + 3, column=base_col, value=vendor.risk_total)
        ws.cell(row=total_row + 4, column=base_col, value=vendor.capex_total)
        ws.cell(row=total_row + 5, column=base_col, value=vendor.grand_total).font = Font(bold=True)
        ws.cell(row=total_row + 6, column=base_col, value=vendor.coverage_pct)

    # Column widths -- best-effort, openpyxl doesn't autosize.
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 44
    for v_idx in range(len(vendor_scores)):
        base_col = 7 + v_idx * 4
        ws.column_dimensions[ws.cell(row=1, column=base_col).column_letter].width = 28
        ws.column_dimensions[ws.cell(row=1, column=base_col + 1).column_letter].width = 10
        ws.column_dimensions[ws.cell(row=1, column=base_col + 2).column_letter].width = 36
        ws.column_dimensions[ws.cell(row=1, column=base_col + 3).column_letter].width = 36


def _append_vendor_columns(ws, vendor_scores: list[VendorScore]) -> None:
    """Append per-vendor columns to an existing template TBE sheet.

    We do *not* touch existing rows; we only add columns to the right
    of the last existing column.
    """
    bold = Font(bold=True)
    start_col = (ws.max_column or 0) + 1
    header_row = 1
    # Find a row that already has values (template headers usually live
    # in row 1 or 2) -- otherwise default to row 1.
    for r in (1, 2, 3, 4):
        if any(ws.cell(row=r, column=c).value for c in range(1, max(2, ws.max_column + 1))):
            header_row = r
            break

    for v_idx, vendor in enumerate(vendor_scores):
        base_col = start_col + v_idx * 4
        ws.cell(row=header_row, column=base_col, value=f"{vendor.vendor} value").font = bold
        ws.cell(row=header_row, column=base_col + 1, value=f"{vendor.vendor} score").font = bold
        ws.cell(row=header_row, column=base_col + 2, value=f"{vendor.vendor} ref").font = bold
        ws.cell(row=header_row, column=base_col + 3, value=f"{vendor.vendor} comment").font = bold
        for offset, (_section, criterion) in enumerate(ROW_PLAN, start=1):
            row = header_row + offset
            line = next((l for l in vendor.lines if l.criterion == criterion), None)
            if line is None:
                ws.cell(row=row, column=base_col, value=NOT_DECLARED)
                ws.cell(row=row, column=base_col + 1, value=0.0)
                ws.cell(row=row, column=base_col + 2, value=f"Ref: {NOT_DECLARED}")
                ws.cell(row=row, column=base_col + 3, value="Row missing.")
            else:
                ws.cell(row=row, column=base_col, value=line.value if not isinstance(line.value, (list, dict)) else str(line.value))
                ws.cell(row=row, column=base_col + 1, value=line.score)
                ws.cell(row=row, column=base_col + 2, value=line.ref)
                ws.cell(row=row, column=base_col + 3, value=line.comment)


def write_master_summary(per_package: dict[str, list[VendorScore]], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Summary"
    ws["A1"] = "Krayavikrayam -- TBE Master Summary"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ("Package", "Vendor", "Design", "Reliability", "Ease", "Risk", "Capex", "Grand total", "Coverage %")
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h).font = Font(bold=True)
    row = 4
    for pkg_key, scores in per_package.items():
        for vs in scores:
            ws.cell(row=row, column=1, value=pkg_key)
            ws.cell(row=row, column=2, value=vs.vendor)
            ws.cell(row=row, column=3, value=vs.design_total)
            ws.cell(row=row, column=4, value=vs.reliability_total)
            ws.cell(row=row, column=5, value=vs.ease_total)
            ws.cell(row=row, column=6, value=vs.risk_total)
            ws.cell(row=row, column=7, value=vs.capex_total)
            ws.cell(row=row, column=8, value=vs.grand_total)
            ws.cell(row=row, column=9, value=vs.coverage_pct)
            row += 1
    for col_letter, width in (("A", 16), ("B", 22), ("C", 12), ("D", 14),
                              ("E", 10), ("F", 10), ("G", 10), ("H", 14), ("I", 14)):
        ws.column_dimensions[col_letter].width = width
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Wrote Master Summary: %s", out_path)
    return out_path


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def fill_for_package(
    package_key: str,
    vendors: list[str],
    evidence_root: Path = EVIDENCE_ROOT,
    out_root: Path = DELIVERABLES_ROOT,
    version_date: str | None = None,
) -> tuple[Path, list[VendorScore]]:
    package = by_key(package_key)
    scores: list[VendorScore] = []
    for vendor in vendors:
        ev = _read_evidence(package, vendor, evidence_root)
        scores.append(score_vendor(package, vendor, ev))
    date = version_date or datetime.utcnow().strftime("%Y%m%d")
    out_path = out_root / f"TBE-{package.key}-v{date}.xlsx"
    write_workbook(package, scores, out_path)
    return out_path, scores


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    parser = argparse.ArgumentParser(description="Fill TBE workbook for a package.")
    parser.add_argument("--package", required=True)
    parser.add_argument("--vendors", required=True, help="Comma-separated vendor names")
    parser.add_argument("--evidence-root", default=str(EVIDENCE_ROOT))
    parser.add_argument("--out-root", default=str(DELIVERABLES_ROOT))
    args = parser.parse_args(argv)

    vendors = [v.strip() for v in args.vendors.split(",") if v.strip()]
    path, _ = fill_for_package(
        package_key=args.package,
        vendors=vendors,
        evidence_root=Path(args.evidence_root),
        out_root=Path(args.out_root),
    )
    print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
