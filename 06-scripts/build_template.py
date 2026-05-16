"""Build a placeholder TBE-UV2.xlsx that mirrors the canonical 4-in-1 layout.

This script is run ONCE to seed 01-templates/TBE-UV2.xlsx so fill_tbe.py has
something to open. Replace the produced file with the real master template the
moment one is available — fill_tbe.py reads cells by header label, not by
hard-coded address, so a structurally-equivalent template will Just Work.

UV-based weight overrides per 00-rules/CLAUDE.md:
  Design Concept (Technical Compliance) = 30
  Ease of Operation & Maintenance / Footprint = 5
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "01-templates" / "TBE-UV2.xlsx"

VENDORS = ["Aster", "ATT", "Labtech", "Zuvay", "Stech", "UFE", "Zenitek", "CME", "Zealwe"]

# (Category, CatPct, CatWt, Subcategory, Weight, IsCapex)
ROWS = [
    ("Capex",      35, 35, "Machine Cost",                                  20, True),
    ("Capex",      35, 35, "Lead time",                                      8, True),
    ("Capex",      35, 35, "Market size / position",                         4, True),
    ("Capex",      35, 35, "Commercial Terms",                               3, True),
    ("Operation",  15, 15, "Ease of Operation & Maintenance(Automation)/Footprint", 5, False),
    ("Operation",  15, 15, "Equipment Availability (uptime)",                2, False),
    ("Operation",  15, 15, "Spare Part Availability & origin",               8, False),
    ("Operation",  15, 15, "Service & Support Lead time",                   14, False),
    ("Technology", 50, 50, "Design Concept (technical compliance)",         30, False),
    ("Technology", 50, 50, "Lessons learnt / responsiveness",                8, False),
    ("Technology", 50, 50, "Technology Enabler (communication protocol / MES / Digital)", 7, False),
    ("Technology", 50, 50, "Technology Maturity (already covered in TS)",    5, False),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CAPEX_FILL  = PatternFill("solid", fgColor="FFF2CC")
OP_FILL     = PatternFill("solid", fgColor="DEEBF7")
TECH_FILL   = PatternFill("solid", fgColor="E2EFDA")


def build() -> None:
    wb = Workbook()

    # --- main sheet ---
    ws = wb.active
    ws.title = "0-TBE"

    ws["A1"] = "TBE — UV-2 Walk-in Chamber (TC + UV, 2-module)"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + 3 * len(VENDORS) + 1)

    ws["A2"] = "Ref RFP: iRIL-MxTR-FRM-GE-001-Rev00 | Source: UV2-comparison-file.xlsx | Date: 2026-05-16"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4 + 3 * len(VENDORS) + 1)

    # column header row (row 4)
    headers = ["Cat %", "Cat Wt", "Category", "Subcategory", "W"]
    for v in VENDORS:
        headers += [f"{v} — Details", f"{v} — Score", f"{v} — SxW"]
    headers += ["Category Owner"]

    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # data rows starting row 5
    for idx, (cat, pct, wt, sub, w, is_capex) in enumerate(ROWS):
        r = 5 + idx
        ws.cell(row=r, column=1, value=pct if idx == 0 or ROWS[idx - 1][0] != cat else None)
        ws.cell(row=r, column=2, value=wt if idx == 0 or ROWS[idx - 1][0] != cat else None)
        ws.cell(row=r, column=3, value=cat if idx == 0 or ROWS[idx - 1][0] != cat else None)
        ws.cell(row=r, column=4, value=sub)
        ws.cell(row=r, column=5, value=w)

        fill = CAPEX_FILL if cat == "Capex" else OP_FILL if cat == "Operation" else TECH_FILL
        for col in range(1, len(headers) + 1):
            cc = ws.cell(row=r, column=col)
            if cc.value is None:
                cc.value = None
            cc.alignment = Alignment(vertical="top", wrap_text=True)
            if col <= 5 or col == len(headers):
                cc.fill = fill

        ws.cell(row=r, column=len(headers),
                value="Project & Sourcing" if cat == "Capex"
                else "Production / Lab Ops" if cat == "Operation"
                else "Module R&D Lab")

    # totals row
    tot_row = 5 + len(ROWS) + 1
    ws.cell(row=tot_row, column=4, value="TOTAL (Weighted, max 1000)").font = Font(bold=True)
    ws.cell(row=tot_row, column=5, value=sum(r[4] for r in ROWS))

    # column widths
    widths = [7, 7, 14, 38, 6] + sum(([40, 8, 8] for _ in VENDORS), []) + [22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 32
    for idx in range(len(ROWS)):
        ws.row_dimensions[5 + idx].height = 48

    # --- protected stub sheets ---
    for name in ("1-comparison", "2-Utilities", "3-Warranty", "4-BOM"):
        s = wb.create_sheet(name)
        s["A1"] = f"PROTECTED — fill_tbe.py MUST NOT write to this sheet ({name})."
        s["A1"].font = Font(italic=True, color="808080")
        s.column_dimensions["A"].width = 80

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)}  ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    build()
