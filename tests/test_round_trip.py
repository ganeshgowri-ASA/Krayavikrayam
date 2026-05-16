"""openpyxl round-trip integrity tests.

Verifies that:
  * a fresh workbook can be written and re-opened cleanly;
  * if a master template exists, ``fill_tbe`` preserves it (no
    regeneration from scratch) -- including a formula cell, a merged-cell
    range, and an embedded image-like binary;
  * the Master Summary workbook round-trips with the expected vendor
    rows.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

from scripts.fill_tbe import (
    fill_for_package,
    load_or_build_workbook,
    write_master_summary,
)
from scripts.packages import by_key
from scripts.scoring import ScoreLine


_PNG_1x1 = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c489"
    "0000000a49444154789c6300010000000500010d0a2db40000000049454e44ae426082"
)


@pytest.fixture
def evidence_dir(tmp_path: Path) -> Path:
    d = tmp_path / "evidence"
    d.mkdir()
    return d


def _write_evidence(evidence_dir: Path, package: str, vendor: str, payload: dict) -> None:
    pkg_dir = evidence_dir / package
    pkg_dir.mkdir(parents=True, exist_ok=True)
    import json
    (pkg_dir / f"{vendor}.json").write_text(json.dumps(payload))


def test_fresh_workbook_round_trips(tmp_path, evidence_dir):
    _write_evidence(evidence_dir, "DH20", "Atlas", {})
    out, scores = fill_for_package(
        package_key="DH20",
        vendors=["Atlas"],
        evidence_root=evidence_dir,
        out_root=tmp_path / "deliverables",
        version_date="20260516",
    )
    assert out.exists()
    wb = load_workbook(out)
    assert "TBE" in wb.sheetnames
    ws = wb["TBE"]
    # Header row at row 4.
    assert ws.cell(row=4, column=1).value == "Section"
    assert ws.cell(row=4, column=2).value == "Criterion"


def test_template_preserves_formula_merged_cells_and_image(tmp_path, evidence_dir, monkeypatch):
    # Build a 'master template' on disk that includes everything we
    # claim to preserve.
    templates_dir = tmp_path / "templates"
    templates_dir.mkdir()
    tpl_path = templates_dir / "DH20.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Master TBE template -- do not delete"
    ws["A2"] = "Sum:"
    ws["B2"] = "=SUM(C1:C5)"
    ws.merge_cells("A4:D4")
    ws["A4"] = "Merged header"
    img_path = tmp_path / "tiny.png"
    img_path.write_bytes(_PNG_1x1)
    ws.add_image(XLImage(str(img_path)), "F1")
    wb.save(tpl_path)

    # Patch the TEMPLATES_ROOT so fill_tbe picks our template up.
    import scripts.fill_tbe as fill_tbe
    monkeypatch.setattr(fill_tbe, "TEMPLATES_ROOT", templates_dir)

    _write_evidence(evidence_dir, "DH20", "Atlas", {})
    out, _ = fill_for_package(
        package_key="DH20",
        vendors=["Atlas"],
        evidence_root=evidence_dir,
        out_root=tmp_path / "deliverables",
        version_date="20260516",
    )

    wb2 = load_workbook(out)
    assert "Instructions" in wb2.sheetnames, "Template sheet must survive"
    ws2 = wb2["Instructions"]
    assert ws2["A1"].value == "Master TBE template -- do not delete"
    # Formula preserved
    assert ws2["B2"].value == "=SUM(C1:C5)"
    # Merged cell range preserved
    merged_ranges = [str(r) for r in ws2.merged_cells.ranges]
    assert "A4:D4" in merged_ranges
    # Image survives -- openpyxl reads images into ws._images.
    assert ws2._images, "Embedded image must survive round trip"


def test_load_or_build_returns_fresh_when_no_template(tmp_path, monkeypatch):
    monkeypatch.setattr("scripts.fill_tbe.TEMPLATES_ROOT", tmp_path / "missing")
    pkg = by_key("DH20")
    wb, from_template = load_or_build_workbook(pkg)
    assert not from_template
    assert isinstance(wb, Workbook)


def test_master_summary_writes_one_row_per_vendor(tmp_path):
    pkg = by_key("DH20")
    fake_scores = [
        type("S", (), dict(vendor="A", design_total=10.0, reliability_total=5.0,
                            ease_total=2.0, risk_total=-1.0, capex_total=0.0,
                            grand_total=16.0, coverage_pct=80.0))(),
        type("S", (), dict(vendor="B", design_total=8.0, reliability_total=4.0,
                            ease_total=1.0, risk_total=0.0, capex_total=0.0,
                            grand_total=13.0, coverage_pct=70.0))(),
    ]
    path = write_master_summary({"DH20": fake_scores}, tmp_path / "master.xlsx")
    wb = load_workbook(path)
    ws = wb["Master Summary"]
    # Header at row 3, data starts row 4.
    assert ws.cell(row=3, column=1).value == "Package"
    assert ws.cell(row=4, column=2).value == "A"
    assert ws.cell(row=5, column=2).value == "B"
    assert ws.cell(row=4, column=8).value == 16.0
